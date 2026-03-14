"""
Amazon ASIN Price Checker — Web Application
Flask app using SP-API for pricing data with background processing and Excel download.
"""

import os
import uuid
import threading
import io
from datetime import datetime

from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from scraper import run_pricing_job

app = Flask(__name__)

# In-memory job store
jobs = {}


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/scrape", methods=["POST"])
def start_scrape():
    """Start a pricing job in the background."""
    data = request.get_json()
    raw_input = data.get("asins", "")

    # Parse ASINs
    asins = []
    for token in raw_input.replace(",", " ").replace("\t", " ").replace("\n", " ").split():
        token = token.strip().upper()
        if token and len(token) >= 5:
            asins.append(token)

    # Deduplicate
    seen = set()
    unique_asins = []
    for a in asins:
        if a not in seen:
            seen.add(a)
            unique_asins.append(a)

    if not unique_asins:
        return jsonify({"error": "No valid ASINs found."}), 400

    job_id = str(uuid.uuid4())[:8]
    job_state = {
        "status": "queued",
        "progress": 0,
        "total": len(unique_asins),
        "results": [],
        "error": None,
        "created_at": datetime.utcnow().isoformat(),
    }
    jobs[job_id] = job_state

    thread = threading.Thread(
        target=run_pricing_job,
        args=(unique_asins, job_state),
        daemon=True,
    )
    thread.start()

    return jsonify({
        "job_id": job_id,
        "total_asins": len(unique_asins),
        "message": f"Fetching prices for {len(unique_asins)} ASINs...",
    })


@app.route("/api/status/<job_id>")
def job_status(job_id):
    """Get the current status of a pricing job."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "total": job["total"],
        "results": job["results"],
        "error": job.get("error"),
    })


@app.route("/api/download/<job_id>")
def download_excel(job_id):
    """Download results as an Excel file."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    if job["status"] != "complete":
        return jsonify({"error": "Job not yet complete"}), 400

    # Build Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon Prices"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    ok_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    err_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    # Headers
    headers = ["#", "ASIN", "List Price", "Buybox Price", "Your Price", "Landed Price", "# Offers", "Buybox Seller", "FBA", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for i, r in enumerate(job["results"], 1):
        row = i + 1
        seller = r.get("buybox_seller") or {}
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=r["asin"]).border = thin_border
        ws.cell(row=row, column=3, value=r["list_price"]).border = thin_border
        ws.cell(row=row, column=4, value=r["buybox_price"]).border = thin_border
        ws.cell(row=row, column=5, value=r["your_price"]).border = thin_border
        ws.cell(row=row, column=6, value=r["landed_price"]).border = thin_border
        ws.cell(row=row, column=7, value=r["num_offers"]).border = thin_border
        ws.cell(row=row, column=8, value=seller.get("seller_id", "N/A")).border = thin_border
        ws.cell(row=row, column=9, value="Yes" if seller.get("is_fba") else "No").border = thin_border
        ws.cell(row=row, column=10, value=r.get("error") or "OK").border = thin_border

        fill = ok_fill if r["status"] == "ok" else err_fill
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = fill

    # Column widths
    widths = [6, 16, 14, 14, 14, 14, 10, 18, 6, 20]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    ws.freeze_panes = "A2"

    # Summary
    sr = len(job["results"]) + 3
    ok_count = sum(1 for r in job["results"] if r["status"] == "ok")
    ws.cell(row=sr, column=1, value="Summary:").font = Font(bold=True)
    ws.cell(row=sr, column=2, value=f"{ok_count}/{len(job['results'])} prices found")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"amazon_prices_{timestamp}.xlsx",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
